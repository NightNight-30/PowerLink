#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成 powerlink 数据库的数据字典 Excel"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

wb = Workbook()

# ========== 样式 ==========
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
title_font = Font(name='微软雅黑', bold=True, size=14, color='1F4E79')
section_font = Font(name='微软雅黑', bold=True, size=12, color='2E75B6')
normal_font = Font(name='微软雅黑', size=10)
thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

headers = ['序号', '字段名', '中文名称', '数据类型', '长度', '是否主键', '是否允许空', '默认值', '来源系统', '原始字段路径', '转换规则', '备注']
col_widths = [18, 25, 18, 12, 8, 10, 10, 20, 10, 30, 24, 36]


def write_sheet(ws, overview_data, fields_data, dir_sheet_name='目录'):
    start_overview_row = 2
    # 动态计算字段区起始行：概览标题(1行) + 概览数据(N行) + 空行(1行) + 字段标题
    start_fields_row = start_overview_row + 1 + len(overview_data) + 1
    # 返回目录链接
    ws[f'A{start_overview_row - 1}'] = '← 返回目录'
    ws[f'A{start_overview_row - 1}'].font = Font(name='微软雅黑', bold=True, size=10, color='2E75B6', underline='single')
    ws[f'A{start_overview_row - 1}'].hyperlink = Hyperlink(ref=f'A{start_overview_row - 1}', location=f'{dir_sheet_name}!A1', display='← 返回目录')

    # 概览区
    ws.merge_cells(f'A{start_overview_row}:L{start_overview_row}')
    ws[f'A{start_overview_row}'] = '数据表概览'
    ws[f'A{start_overview_row}'].font = title_font
    ws[f'A{start_overview_row}'].alignment = left_align
    overview_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    ws[f'A{start_overview_row}'].fill = overview_fill
    ws[f'A{start_overview_row}'].border = thin_border
    for i, (k, v) in enumerate(overview_data):
        r = start_overview_row + 1 + i
        key_cell = ws[f'A{r}']
        key_cell.value = k
        key_cell.font = Font(name='微软雅黑', bold=True, size=10, color='1F4E79')
        key_cell.alignment = left_align
        key_cell.border = thin_border
        key_cell.fill = overview_fill
        val_cell = ws[f'B{r}']
        val_cell.value = v
        val_cell.font = normal_font
        val_cell.alignment = left_align
        val_cell.border = thin_border
        # 合并B~L列，让值完整显示
        ws.merge_cells(f'B{r}:L{r}')

    # 字段区标题
    title_row = start_fields_row
    ws.merge_cells(f'A{title_row}:L{title_row}')
    ws[f'A{title_row}'] = '字段明细'
    ws[f'A{title_row}'].font = section_font
    ws[f'A{title_row}'].alignment = left_align
    ws[f'A{title_row}'].fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    ws[f'A{title_row}'].border = thin_border

    # 表头
    header_row = title_row + 1
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 数据行
    for row_idx, field in enumerate(fields_data):
        r = header_row + 1 + row_idx
        for col_idx, val in enumerate(field, 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.value = val
            cell.font = normal_font
            cell.alignment = left_align if col_idx > 3 else center_align
            cell.border = thin_border
            if row_idx % 2 == 1:
                cell.fill = alt_fill

    # 列宽
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结窗格：固定表头行，方便滚动浏览字段
    ws.freeze_panes = f'A{header_row + 1}'


# ========== Sheet1: api_call_record ==========
ws1 = wb.active
ws1.title = 'api_call_record'

overview1 = [
    ('数据库', 'powerlink'), ('表名', 'api_call_record'), ('表描述', '三方接口调用记录表'),
    ('引擎', 'InnoDB'), ('字符集', 'utf8mb4'), ('所属系统', '天眼查数据接入'),
    ('负责人', ''), ('创建日期', '2026-05-14'),
]

fields1 = [
    (1, 'id', '主键ID', 'BIGINT', '', 'Y', 'N', '自增', '内部', '-', '-', '自增主键'),
    (2, 'interface_name', '接口名', 'VARCHAR', '32', 'N', 'N', '', '内部', '-', '-', '标识819/967/971/973等接口'),
    (3, 'call_datetime', '调用日期时间', 'DATETIME', '', 'N', 'N', '', '内部', '-', '-', 'API调用发生的时间'),
    (4, 'input_param', '入参(公司名)', 'VARCHAR', '200', 'N', 'N', '', '内部', '-', '-', '搜索关键字/公司名'),
    (5, 'status_code', '状态码', 'INT', '', 'N', 'N', '', '天眼查/邓白氏', 'result.error_code / code', '[枚举:ENUM_DNB_STATUS_CODE]对邓白氏接口', '0=成功,负数=异常,正数=业务错误码; 邓白氏code=1004等见枚举字典'),
    (6, 'output_result', '出参结果', 'JSON', '', 'N', 'Y', '', '天眼查', 'response body', '原始JSON完整存储', '成功时为API完整响应JSON，失败时为错误详情JSON'),
    (7, 'create_time', '创建时间', 'DATETIME', '', 'N', 'Y', 'CURRENT_TIMESTAMP', '内部', '-', '-', '记录入库时间'),
]

write_sheet(ws1, overview1, fields1)

# ========== Sheet2: company_819_info ==========
ws2 = wb.create_sheet('company_819_info')

overview2 = [
    ('数据库', 'powerlink'), ('表名', 'company_819_info'), ('表描述', '企业基本信息(819接口)'),
    ('引擎', 'InnoDB'), ('字符集', 'utf8mb4'), ('所属系统', '天眼查数据接入'),
    ('唯一约束', 'company_name'), ('创建日期', '2026-05-14'),
    ('解析规则说明', 'Array+child String→逗号分隔字符串; Object+多KV→展开为独立列; Object+可能多条→JSON字符串+提取total; Number时间戳→DATETIME'),
]

fields2 = [
    # 基础标识
    (1, 'id', '主键ID', 'BIGINT', '', 'Y', 'N', '自增', '内部', '-', '-', '自增主键'),
    (2, 'api_record_id', 'API调用记录ID', 'BIGINT', '', 'N', 'Y', '', '内部', 'api_call_record.id', '-', '关联api_call_record表，可追溯原始API调用'),
    (3, 'data_create_time', '数据创建时间', 'DATETIME', '', 'N', 'Y', 'CURRENT_TIMESTAMP', '内部', '-', '-', '解析入库时间'),
    (4, 'company_name', '公司名(搜索关键字)', 'VARCHAR', '200', 'UK', 'N', '', '内部', '-', '-', '搜索关键字/公司名，唯一约束'),
    (5, 'company_id', '企业ID', 'BIGINT', '', 'N', 'Y', '', '天眼查', 'result.id', '-', '天眼查企业唯一ID'),

    # 法人/类型
    (6, 'legal_person_type', '法人类型', 'INT', '', 'N', 'Y', '', '天眼查', 'result.type', '-', '1=自然人，2=公司 [枚举:ENUM_LEGAL_PERSON_TYPE]'),
    (7, 'reg_status', '经营状态', 'VARCHAR', '31', 'N', 'Y', '', '天眼查', 'result.regStatus', '[枚举:ENUM_REG_STATUS(推断)]', '如:存续/注销/吊销等'),
    (8, 'legal_person_name', '法定代表人', 'VARCHAR', '120', 'N', 'Y', '', '天眼查', 'result.legalPersonName', '-', ''),
    (9, 'company_org_type', '企业类型', 'VARCHAR', '127', 'N', 'Y', '', '天眼查', 'result.companyOrgType', '-', '如:其他股份有限公司(上市)'),
    (10, 'is_micro_ent', '是否小微企业', 'INT', '', 'N', 'Y', '', '天眼查', 'result.isMicroEnt', '[枚举:ENUM_IS_MICRO_ENT]', '0=否,1=是'),

    # 资本
    (11, 'reg_capital', '注册资本(含单位)', 'VARCHAR', '50', 'N', 'Y', '', '天眼查', 'result.regCapital', '-', '如:730710.9252万人民币'),
    (12, 'reg_capital_currency', '注册资本币种', 'VARCHAR', '10', 'N', 'Y', '', '天眼查', 'result.regCapitalCurrency', '-', '人民币/美元/欧元等 [枚举:ENUM_CURRENCY(推断)]'),
    (13, 'paid_capital', '实缴资本(含单位)', 'VARCHAR', '50', 'N', 'Y', '', '天眼查', 'result.actualCapital', '字段映射: actualCapital→paid_capital', '如:183130.690104万人民币'),
    (14, 'paid_capital_currency', '实缴资本币种', 'VARCHAR', '10', 'N', 'Y', '', '天眼查', 'result.actualCapitalCurrency', '字段映射: actualCapitalCurrency→paid_capital', '人民币/美元/欧元等 [枚举:ENUM_CURRENCY(推断)]'),

    # 时间(时间戳→datetime)
    (15, 'est_date', '成立日期', 'DATETIME', '', 'N', 'Y', '', '天眼查', 'result.estiblishTime', '时间戳→datetime: ≥1e10为毫秒÷1000', '字段映射: estiblishTime→est_date'),
    (16, 'from_date', '经营开始时间', 'DATETIME', '', 'N', 'Y', '', '天眼查', 'result.fromTime', '时间戳→datetime: ≥1e10为毫秒÷1000', '字段映射: fromTime→from_date'),
    (17, 'to_date', '经营结束时间', 'DATETIME', '', 'N', 'Y', '', '天眼查', 'result.toTime', '时间戳→datetime: ≥1e10为毫秒÷1000', 'null表示无固定期限'),
    (18, 'approval_date', '核准时间', 'DATETIME', '', 'N', 'Y', '', '天眼查', 'result.approvedTime', '时间戳→datetime: ≥1e10为毫秒÷1000', '字段映射: approvedTime→approval_date'),
    (19, 'cancel_date', '注销日期', 'DATETIME', '', 'N', 'Y', '', '天眼查', 'result.cancelDate', '时间戳→datetime: ≥1e10为毫秒÷1000', ''),
    (20, 'revoke_date', '吊销日期', 'DATETIME', '', 'N', 'Y', '', '天眼查', 'result.revokeDate', '时间戳→datetime: ≥1e10为毫秒÷1000', ''),
    (21, 'cancel_reason', '注销原因', 'VARCHAR', '500', 'N', 'Y', '', '天眼查', 'result.cancelReason', '-', ''),
    (22, 'revoke_reason', '吊销原因', 'VARCHAR', '500', 'N', 'Y', '', '天眼查', 'result.revokeReason', '-', ''),
    (23, 'update_time', '更新时间', 'DATETIME', '', 'N', 'Y', '', '天眼查', 'result.updateTimes', '时间戳→datetime: ≥1e10为毫秒÷1000', '字段映射: updateTimes→update_time'),

    # 证件编码
    (24, 'social_credit_code', '统一社会信用代码', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.creditCode', '字段映射: creditCode→social_credit_code', ''),
    (25, 'org_code', '组织机构代码', 'VARCHAR', '31', 'N', 'Y', '', '天眼查', 'result.orgNumber', '字段映射: orgNumber→org_code', ''),
    (26, 'tax_number', '纳税人识别号', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.taxNumber', '-', ''),
    (27, 'reg_number', '注册号', 'VARCHAR', '31', 'N', 'Y', '', '天眼查', 'result.regNumber', '-', ''),
    (28, 'brn_number', '商业登记号', 'VARCHAR', '50', 'N', 'Y', '', '天眼查', 'result.BRNNumber', '-', 'BRN缩写保留'),

    # 地理
    (29, 'province_short', '省份简称', 'VARCHAR', '31', 'N', 'Y', '', '天眼查', 'result.base', '字段映射: base→province_short', '如:gd=广东'),
    (30, 'city', '市', 'VARCHAR', '20', 'N', 'Y', '', '天眼查', 'result.city', '-', ''),
    (31, 'district', '区', 'VARCHAR', '20', 'N', 'Y', '', '天眼查', 'result.district', '-', ''),
    (32, 'district_code', '行政区划代码', 'VARCHAR', '20', 'N', 'Y', '', '天眼查', 'result.districtCode', '-', ''),
    (33, 'reg_location', '注册地址', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.regLocation', '-', ''),
    (34, 'reg_location_half_width', '注册地址(半角)', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.regLocationHalfWidth', '-', '半角字符版本'),
    (35, 'reg_institute', '登记机关', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.regInstitute', '-', ''),
    (36, 'economic_function_zone1', '经济功能区1', 'VARCHAR', '20', 'N', 'Y', '', '天眼查', 'result.economicFunctionZone1', '-', ''),
    (37, 'economic_function_zone2', '经济功能区2', 'VARCHAR', '20', 'N', 'Y', '', '天眼查', 'result.economicFunctionZone2', '-', ''),
    (38, 'above_scale', '是否规模以上', 'VARCHAR', '10', 'N', 'Y', '', '天眼查', 'result.aboveScale', '-', '如:规模以上工业'),

    # 经营范围/行业
    (39, 'business_scope', '经营范围', 'TEXT', '', 'N', 'Y', '', '天眼查', 'result.businessScope', '-', ''),
    (40, 'industry', '行业', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.industry', '-', ''),

    # 名称/别名
    (41, 'company_alias', '简称', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.alias', '字段映射: alias→company_alias', '企业简称/别名'),
    (42, 'used_bond_name', '股票曾用名', 'VARCHAR', '50', 'N', 'Y', '', '天眼查', 'result.usedBondName', '-', ''),
    (43, 'property3', '英文名', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.property3', '-', 'property3缩写保留'),

    # 联系方式
    (44, 'email', '邮箱(单个)', 'VARCHAR', '1024', 'N', 'Y', '', '天眼查', 'result.email', '-', 'API返回的第一个邮箱'),
    (45, 'email_list', '全部邮箱', 'TEXT', '', 'N', 'Y', '', '天眼查', 'result.emailList', 'Array+child String → 逗号分隔字符串', '如: a@b.com,c@d.com,d@e.com'),
    (46, 'phone_number', '企业联系方式', 'VARCHAR', '1024', 'N', 'Y', '', '天眼查', 'result.phoneNumber', '-', ''),
    (47, 'website_list', '网址', 'TEXT', '', 'N', 'Y', '', '天眼查', 'result.websiteList', '-', ''),

    # 曾用名
    (48, 'history_names', '曾用名', 'TEXT', '', 'N', 'Y', '', '天眼查', 'result.historyNameList', 'Array+child String → 逗号分隔字符串', '忽略historyNames(分号分隔)，取historyNameList(JSON数组)'),

    # 股票
    (49, 'bond_name', '股票名', 'VARCHAR', '20', 'N', 'Y', '', '天眼查', 'result.bondName', '-', ''),
    (50, 'bond_num', '股票号', 'VARCHAR', '20', 'N', 'Y', '', '天眼查', 'result.bondNum', '-', ''),
    (51, 'bond_type', '股票类型', 'VARCHAR', '31', 'N', 'Y', '', '天眼查', 'result.bondType', '-', '如:A股'),

    # 评分/规模
    (52, 'tags', '企业标签', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.tags', '-', ''),
    (53, 'staff_num_range', '人员规模', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.staffNumRange', '-', '如:5000-9999人'),
    (54, 'social_staff_num', '参保人数', 'INT', '', 'N', 'Y', '', '天眼查', 'result.socialStaffNum', '-', ''),
    (55, 'percentile_score', '企业评分(万分制)', 'INT', '', 'N', 'Y', '', '天眼查', 'result.percentileScore', '-', ''),

    # staffList(Object+可能多条)
    (56, 'staff_list_total', '主要人员总数', 'INT', '', 'N', 'Y', '', '天眼查', 'result.staffList.total', 'Object提取total字段', ''),
    (57, 'staff_list_json', '主要人员列表', 'TEXT', '', 'N', 'Y', '', '天眼查', 'result.staffList.result', 'Object+可能多条 → JSON字符串', 'JSON数组，每项含name/id/type/typeJoin'),

    # industryAll(Object+多KV展开)
    (58, 'industry_all_category', '国民经济行业分类-门类', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.industryAll.category', 'Object+多KV展开', '如:制造业'),
    (59, 'industry_all_category_big', '国民经济行业分类-大类', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.industryAll.categoryBig', 'Object+多KV展开', '如:计算机、通信和其他电子设备制造业'),
    (60, 'industry_all_category_middle', '国民经济行业分类-中类', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.industryAll.categoryMiddle', 'Object+多KV展开', '如:其他电子设备制造'),
    (61, 'industry_all_category_small', '国民经济行业分类-小类', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.industryAll.categorySmall', 'Object+多KV展开', ''),
    (62, 'industry_all_category_code_first', '国民经济行业分类-门类代码', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.industryAll.categoryCodeFirst', 'Object+多KV展开', '如:C'),
    (63, 'industry_all_category_code_second', '国民经济行业分类-大类代码', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.industryAll.categoryCodeSecond', 'Object+多KV展开', '如:39'),
    (64, 'industry_all_category_code_third', '国民经济行业分类-中类代码', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.industryAll.categoryCodeThird', 'Object+多KV展开', '如:399'),
    (65, 'industry_all_category_code_fourth', '国民经济行业分类-小类代码', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.industryAll.categoryCodeFourth', 'Object+多KV展开', '如:3990'),
]

write_sheet(ws2, overview2, fields2)

# ========== Sheet3: customer_info ==========
ws3 = wb.create_sheet('customer_info')

overview3 = [
    ('数据库', 'powerlink'), ('表名', 'customer_info'), ('表描述', '客户公司列表'),
    ('引擎', 'InnoDB'), ('字符集', 'utf8mb4'), ('所属系统', '天眼查数据接入'),
    ('负责人', ''), ('创建日期', '2026-05-14'),
]

fields3 = [
    (1, 'id', '主键ID', 'BIGINT', '', 'Y', 'N', '自增', '内部', '-', '-', '自增主键'),
    (2, 'customer_name', '公司名', 'VARCHAR', '200', 'N', 'N', '', '内部', '-', '-', '客户公司名称，用于API搜索关键字'),
    (3, 'create_time', '创建时间', 'DATETIME', '', 'N', 'Y', 'CURRENT_TIMESTAMP', '内部', '-', '-', '记录入库时间'),
]

write_sheet(ws3, overview3, fields3)

# ========== Sheet4: company_1058_risk_info ==========
ws4 = wb.create_sheet('company_1058_risk_info')

overview4 = [
    ('数据库', 'powerlink'), ('表名', 'company_1058_risk_info'), ('表描述', '企业天眼风险(1058接口)'),
    ('引擎', 'InnoDB'), ('字符集', 'utf8mb4'), ('所属系统', '天眼查数据接入'),
    ('数据关系', '1:N(1公司→N条风险记录)'), ('创建日期', '2026-05-14'),
    ('解析规则说明', '3层嵌套展平(riskList→list→list); DELETE旧数据+INSERT新数据; main_company_name来自搜索入参'),
]

fields4 = [
    (1, 'id', '主键ID', 'BIGINT', '', 'Y', 'N', '自增', '内部', '-', '-', '自增主键'),
    (2, 'api_record_id', 'API调用记录ID', 'BIGINT', '', 'N', 'Y', '', '内部', 'api_call_record.id', '-', '关联api_call_record表，可追溯原始API调用'),
    (3, 'data_create_time', '数据创建时间', 'DATETIME', '', 'N', 'Y', 'CURRENT_TIMESTAMP', '内部', '-', '-', '解析入库时间'),
    (4, 'main_company_name', '主公司名(搜索关键字)', 'VARCHAR', '200', 'N', 'N', '', '内部', '-', '-', '来自搜索入参(input_param)，非API返回'),
    (5, 'risk_level', '风险等级', 'VARCHAR', '50', 'N', 'Y', '', '天眼查', 'result.riskLevel', '-', '顶层风险等级'),
    (6, 'risk_category_count', '风险类别下的条数', 'INT', '', 'N', 'Y', '', '天眼查', 'result.riskList[].count', '3层嵌套展平meta字段', '该风险类别下的风险条数'),
    (7, 'risk_category_name', '风险类别名', 'VARCHAR', '50', 'N', 'Y', '', '天眼查', 'result.riskList[].name', '3层嵌套展平meta字段', '自身风险/周边风险/历史风险/预警提醒'),
    (8, 'risk_type_total', '风险类型下的条数', 'INT', '', 'N', 'Y', '', '天眼查', 'result.riskList[].list[].total', '3层嵌套展平meta字段', '该风险类型组下的条数'),
    (9, 'risk_type_tag', '风险标签', 'VARCHAR', '50', 'N', 'Y', '', '天眼查', 'result.riskList[].list[].tag', '[枚举:ENUM_RISK_TAG]', '高风险/警示/提示信息'),
    (10, 'company_id', '涉及公司ID', 'BIGINT', '', 'N', 'Y', '', '天眼查', 'result.riskList[].list[].list[].companyId', '-', '可空，风险涉及的公司ID'),
    (11, 'company_name', '涉及公司名', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.riskList[].list[].list[].companyName', '空字符串→NULL', '可空，风险涉及的公司名'),
    (15, 'risk_type', '风险类型码', 'INT', '', 'N', 'Y', '', '天眼查', 'result.riskList[].list[].list[].type', '[枚举:ENUM_RISK_TYPE]', '约50个值,见枚举字典'),
    (13, 'risk_count', '风险数量', 'INT', '', 'N', 'Y', '', '天眼查', 'result.riskList[].list[].list[].riskCount', '-', ''),
    (14, 'risk_title', '风险描述', 'VARCHAR', '500', 'N', 'Y', '', '天眼查', 'result.riskList[].list[].list[].title', '-', '风险详情描述'),
    (15, 'risk_type', '风险类型码', 'INT', '', 'N', 'Y', '', '天眼查', 'result.riskList[].list[].list[].type', '-', '风险类型编码'),
    (16, 'risk_desc', '风险简述', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.riskList[].list[].list[].desc', '-', '风险简要分类描述'),
]

write_sheet(ws4, overview4, fields4)

# ========== Sheet5: company_822_change_info ==========
ws5 = wb.create_sheet('company_822_change_info')

overview5 = [
    ('数据库', 'powerlink'), ('表名', 'company_822_change_info'), ('表描述', '变更记录(822接口)'),
    ('引擎', 'InnoDB'), ('字符集', 'utf8mb4'), ('所属系统', '天眼查数据接入'),
    ('数据关系', '1:N(1公司→N条变更记录)'), ('创建日期', '2026-05-15'),
    ('解析规则说明', '2层展平(result.items数组); DELETE旧数据+INSERT新数据; company_name来自搜索入参'),
]

fields5 = [
    (1, 'id', '主键ID', 'BIGINT', '', 'Y', 'N', '自增', '内部', '-', '-', '自增主键'),
    (2, 'api_record_id', 'API调用记录ID', 'BIGINT', '', 'N', 'Y', '', '内部', 'api_call_record.id', '-', '关联api_call_record表'),
    (3, 'data_create_time', '数据创建时间', 'DATETIME', '', 'N', 'Y', 'CURRENT_TIMESTAMP', '内部', '-', '-', '解析入库时间'),
    (4, 'company_name', '主公司名(搜索关键字)', 'VARCHAR', '200', 'N', 'N', '', '内部', '-', '-', '来自搜索入参(input_param)，非API返回'),
    (5, 'total', '变更记录总数', 'INT', '', 'N', 'Y', '', '天眼查', 'result.total', '2层展平meta字段', '变更记录总条数'),
    (6, 'change_item', '变更项名称', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.items[].changeItem', '驼峰→下划线', '如: 经营范围变更'),
    (7, 'content_before', '变更前内容', 'TEXT', '', 'N', 'Y', '', '天眼查', 'result.items[].contentBefore', '空字符串→NULL', ''),
    (8, 'content_after', '变更后内容', 'TEXT', '', 'N', 'Y', '', '天眼查', 'result.items[].contentAfter', '空字符串→NULL', ''),
    (9, 'change_time', '变更时间', 'VARCHAR', '20', 'N', 'Y', '', '天眼查', 'result.items[].changeTime', '-', '日期字符串，如2025-03-26'),
    (10, 'create_time', '记录创建时间', 'VARCHAR', '20', 'N', 'Y', '', '天眼查', 'result.items[].createTime', '-', '日期字符串，如2025-03-27'),
]

write_sheet(ws5, overview5, fields5)

# ========== Sheet6: company_854_stock_info ==========
ws6 = wb.create_sheet('company_854_stock_info')

overview6 = [
    ('数据库', 'powerlink'), ('表名', 'company_854_stock_info'), ('表描述', '上市公司企业简介(854接口)'),
    ('引擎', 'InnoDB'), ('字符集', 'utf8mb4'), ('所属系统', '天眼查数据接入'),
    ('唯一约束', 'company_name'), ('创建日期', '2026-05-15'),
    ('解析规则说明', '1:1关系, ON DUPLICATE KEY UPDATE; 4个Object字段展开为type/name/id 3列; 非上市公司result为空→跳过; company_name来自搜索入参'),
]

fields6 = [
    (1, 'id', '主键ID', 'BIGINT', '', 'Y', 'N', '自增', '内部', '-', '-', '自增主键'),
    (2, 'api_record_id', 'API调用记录ID', 'BIGINT', '', 'N', 'Y', '', '内部', 'api_call_record.id', '-', '关联api_call_record表'),
    (3, 'data_create_time', '数据创建时间', 'DATETIME', '', 'N', 'Y', 'CURRENT_TIMESTAMP', '内部', '-', '-', '解析入库时间'),
    (4, 'company_name', '主公司名(搜索关键字)', 'VARCHAR', '200', 'UK', 'N', '', '内部', '-', '-', '来自搜索入参(input_param)，非API返回'),
    (5, 'area', '区域', 'LONGTEXT', '', 'N', 'Y', '', '天眼查', 'result.area', '空字符串→NULL', ''),
    (6, 'website', '网址', 'LONGTEXT', '', 'N', 'Y', '', '天眼查', 'result.website', '空字符串→NULL', ''),
    (7, 'stock_code', '股票代码', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.code', '字段映射: code→stock_code', '如: 002600'),
    (8, 'address', '地址', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.address', '空字符串→NULL', ''),
    (9, 'gm_type', '总经理类型', 'INT', '', 'N', 'Y', '', '天眼查', 'result.generalManager.cType', 'Object展开: generalManager→gm_type/gm_name/gm_id; [枚举:ENUM_SUBJECT_TYPE]', '1=公司,2=人'),
    (10, 'gm_name', '总经理姓名', 'VARCHAR', '120', 'N', 'Y', '', '天眼查', 'result.generalManager.name', 'Object展开', ''),
    (11, 'gm_id', '总经理ID', 'BIGINT', '', 'N', 'Y', '', '天眼查', 'result.generalManager.id', 'Object展开; id="0"→NULL', ''),
    (12, 'stock_company_name', 'API返回公司名', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.companyName', '字段映射: companyName→stock_company_name', '区别于入参company_name'),
    (13, 'employees_num', '员工人数', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.employeesNum', '驼峰→下划线', '如: 100434'),
    (14, 'main_business', '主营业务', 'LONGTEXT', '', 'N', 'Y', '', '天眼查', 'result.mainBusiness', '驼峰→下划线; 空字符串→NULL', ''),
    (15, 'mobile', '电话', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.mobile', '空字符串→NULL', '如: 86-0755-25863893'),
    (16, 'chairman_type', '董事长类型', 'INT', '', 'N', 'Y', '', '天眼查', 'result.chairman.cType', 'Object展开; [枚举:ENUM_SUBJECT_TYPE]', '1=公司,2=人'),
    (17, 'chairman_name', '董事长姓名', 'VARCHAR', '120', 'N', 'Y', '', '天眼查', 'result.chairman.name', 'Object展开', ''),
    (18, 'chairman_id', '董事长ID', 'BIGINT', '', 'N', 'Y', '', '天眼查', 'result.chairman.id', 'Object展开; id="0"→NULL', ''),
    (19, 'industry', '行业', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.industry', '空字符串→NULL', '如: 电子 — 消费电子'),
    (20, 'product_name', '产品名称', 'LONGTEXT', '', 'N', 'Y', '', '天眼查', 'result.productName', '驼峰→下划线; 空字符串→NULL', ''),
    (21, 'secretary_type', '董秘类型', 'INT', '', 'N', 'Y', '', '天眼查', 'result.secretaries.cType', 'Object展开; [枚举:ENUM_SUBJECT_TYPE]', '1=公司,2=人'),
    (22, 'secretary_name', '董秘姓名', 'VARCHAR', '120', 'N', 'Y', '', '天眼查', 'result.secretaries.name', 'Object展开', ''),
    (23, 'secretary_id', '董秘ID', 'BIGINT', '', 'N', 'Y', '', '天眼查', 'result.secretaries.id', 'Object展开; id="0"→NULL', ''),
    (24, 'actual_controller', '实际控制人', 'LONGTEXT', '', 'N', 'Y', '', '天眼查', 'result.actualController', '驼峰→下划线; 空字符串→NULL', ''),
    (25, 'controlling_shareholder', '控股股东', 'LONGTEXT', '', 'N', 'Y', '', '天眼查', 'result.controllingShareholder', '驼峰→下划线; 空字符串→NULL', ''),
    (26, 'eng_name', '英文名', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.engName', '驼峰→下划线', ''),
    (27, 'registered_capital', '注册资本', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.registeredCapital', '驼峰→下划线; 空字符串→NULL', '如: 730710.925万人民币'),
    (28, 'postalcode', '邮编', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.postalcode', '-', '已是下划线命名'),
    (29, 'legal_person_type', '法人类型', 'INT', '', 'N', 'Y', '', '天眼查', 'result.legal.cType', 'Object展开; [枚举:ENUM_SUBJECT_TYPE]', '1=公司,2=人'),
    (30, 'legal_person_name', '法人姓名', 'VARCHAR', '120', 'N', 'Y', '', '天眼查', 'result.legal.name', 'Object展开', ''),
    (31, 'legal_person_id', '法人ID', 'BIGINT', '', 'N', 'Y', '', '天眼查', 'result.legal.id', 'Object展开; id="0"→NULL', ''),
    (32, 'listed_name', '上市公司简称', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.name', '字段映射: name→listed_name', '区别于company_name(搜索入参)'),
    (33, 'fax', '传真', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.fax', '空字符串→NULL', ''),
    (34, 'used_name', '曾用名', 'VARCHAR', '255', 'N', 'Y', '', '天眼查', 'result.usedName', '驼峰→下划线; 空字符串→NULL', ''),
    (35, 'final_controller', '最终控制人', 'LONGTEXT', '', 'N', 'Y', '', '天眼查', 'result.finalController', '驼峰→下划线; 空字符串→NULL', ''),
    (36, 'introduction', '简介', 'TEXT', '', 'N', 'Y', '', '天眼查', 'result.introduction', '空字符串→NULL', ''),
]

write_sheet(ws6, overview6, fields6)

# ========== Sheet7: company_1168_org_type_info ==========
ws7 = wb.create_sheet('company_1168_org_type_info')

overview7 = [
    ('数据库', 'powerlink'), ('表名', 'company_1168_org_type_info'), ('表描述', '组织机构类型(1168接口)'),
    ('引擎', 'InnoDB'), ('字符集', 'utf8mb4'), ('所属系统', '天眼查数据接入'),
    ('唯一约束', 'company_name'), ('创建日期', '2026-05-15'),
    ('解析规则说明', '1:1关系, ON DUPLICATE KEY UPDATE; orgTypes/economyTypes数组→逗号分隔level1/level2列'),
]

fields7 = [
    (1, 'id', '主键ID', 'BIGINT', '', 'Y', 'N', '自增', '内部', '-', '-', '自增主键'),
    (2, 'api_record_id', 'API调用记录ID', 'BIGINT', '', 'N', 'Y', '', '内部', 'api_call_record.id', '-', '关联api_call_record表'),
    (3, 'data_create_time', '数据创建时间', 'DATETIME', '', 'N', 'Y', 'CURRENT_TIMESTAMP', '内部', '-', '-', '解析入库时间'),
    (4, 'company_name', '主公司名(搜索关键字)', 'VARCHAR', '200', 'UK', 'N', '', '内部', '-', '-', '来自搜索入参(input_param)，非API返回'),
    (5, 'org_type_level1', '一级机构类型', 'TEXT', '', 'N', 'Y', '', '天眼查', 'result.orgTypes[].level1', 'Array+child Object→逗号分隔', '如: 企业,企业'),
    (6, 'org_type_level2', '二级机构类型', 'TEXT', '', 'N', 'Y', '', '天眼查', 'result.orgTypes[].level2', 'Array+child Object→逗号分隔', '如: 股份有限公司'),
    (7, 'economy_type_level1', '一级经济类型', 'TEXT', '', 'N', 'Y', '', '天眼查', 'result.economyTypes[].level1', 'Array+child Object→逗号分隔', '如: 民营企业'),
    (8, 'economy_type_level2', '二级经济类型', 'TEXT', '', 'N', 'Y', '', '天眼查', 'result.economyTypes[].level2', 'Array+child Object→逗号分隔; null→NULL', '如: 央企，可空'),
]

write_sheet(ws7, overview7, fields7)

# ========== Sheet8: company_1149_scale_info ==========
ws8 = wb.create_sheet('company_1149_scale_info')

overview8 = [
    ('数据库', 'powerlink'), ('表名', 'company_1149_scale_info'), ('表描述', '企业规模(1149接口)'),
    ('引擎', 'InnoDB'), ('字符集', 'utf8mb4'), ('所属系统', '天眼查数据接入'),
    ('唯一约束', 'company_name'), ('创建日期', '2026-05-15'),
    ('解析规则说明', '1:1关系, ON DUPLICATE KEY UPDATE; result为简单字符串(如"大型")'),
]

fields8 = [
    (1, 'id', '主键ID', 'BIGINT', '', 'Y', 'N', '自增', '内部', '-', '-', '自增主键'),
    (2, 'api_record_id', 'API调用记录ID', 'BIGINT', '', 'N', 'Y', '', '内部', 'api_call_record.id', '-', '关联api_call_record表'),
    (3, 'data_create_time', '数据创建时间', 'DATETIME', '', 'N', 'Y', 'CURRENT_TIMESTAMP', '内部', '-', '-', '解析入库时间'),
    (4, 'company_name', '主公司名(搜索关键字)', 'VARCHAR', '200', 'UK', 'N', '', '内部', '-', '-', '来自搜索入参(input_param)，非API返回'),
    (5, 'company_scale', '企业规模', 'VARCHAR', '50', 'N', 'Y', '', '天眼查', 'result', 'result直接为字符串; [枚举:ENUM_COMPANY_SCALE(推断)]', '如:大型/中型/小型/微型'),
]

write_sheet(ws8, overview8, fields8)

# ========== Sheet9: company_967_main_index_info ==========
ws9 = wb.create_sheet('company_967_main_index_info')

overview9 = [
    ('数据库', 'powerlink'), ('表名', 'company_967_main_index_info'), ('表描述', '主要指标-年度(967接口)'),
    ('引擎', 'InnoDB'), ('字符集', 'utf8mb4'), ('所属系统', '天眼查数据接入'),
    ('数据关系', '1:N(1公司→N个年度)'), ('创建日期', '2026-05-15'),
    ('解析规则说明', 'result数组展平,每年度一行; ~28个DECIMAL字段; DELETE旧数据+INSERT新数据; 非上市公司返回300000→step2天然跳过'),
]

fields9 = [
    (1, 'id', '主键ID', 'BIGINT', '', 'Y', 'N', '自增', '内部', '-', '-', '自增主键'),
    (2, 'api_record_id', 'API调用记录ID', 'BIGINT', '', 'N', 'Y', '', '内部', 'api_call_record.id', '-', '关联api_call_record表'),
    (3, 'data_create_time', '数据创建时间', 'DATETIME', '', 'N', 'Y', 'CURRENT_TIMESTAMP', '内部', '-', '-', '解析入库时间'),
    (4, 'company_name', '主公司名(搜索关键字)', 'VARCHAR', '200', 'N', 'N', '', '内部', '-', '-', '来自搜索入参(input_param)，非API返回'),
    (5, 'show_year', '年份', 'VARCHAR', '32', 'N', 'Y', '', '天眼查', 'result[].showYear', '驼峰→下划线', '如: 2023, 2024(Q1)'),
    (6, 'crfgsasr_to_revenue', '销售现金流/营业收入', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].crfgsasr_to_revenue', '-', ''),
    (7, 'np_atsopc_nrgal_yoy', '扣非净利润同比增长(%)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].np_atsopc_nrgal_yoy', '-', ''),
    (8, 'asset_liab_ratio', '资产负债率(%)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].asset_liab_ratio', '-', ''),
    (9, 'op_to_revenue', '营业利润/营业收入(%)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].op_to_revenue', '-', ''),
    (10, 'revenue_yoy', '营业总收入同比增长(%)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].revenue_yoy', '-', ''),
    (11, 'net_profit_atsopc_yoy', '归属净利润同比增长(%)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].net_profit_atsopc_yoy', '-', ''),
    (12, 'receivable_turnover_days', '应收账款周转天数(天)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].receivable_turnover_days', '-', ''),
    (13, 'current_ratio', '流动比率', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].current_ratio', '-', ''),
    (14, 'operate_cash_flow_ps', '每股经营现金流(元)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].operate_cash_flow_ps', '-', ''),
    (15, 'gross_selling_rate', '毛利率(%)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].gross_selling_rate', '-', ''),
    (16, 'current_liab_to_total_liab', '流动负债/总负债(%)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].current_liab_to_total_liab', '-', ''),
    (17, 'quick_ratio', '速动比率', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].quick_ratio', '-', ''),
    (18, 'fully_dlt_roe', '摊薄净资产收益率(%)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].fully_dlt_roe', '-', ''),
    (19, 'tax_rate', '实际税率(%)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].tax_rate', '-', ''),
    (20, 'net_interest_of_total_assets', '摊薄总资产收益率(%)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].net_interest_of_total_assets', '-', ''),
    (21, 'operating_total_revenue_lrr_sq', '营业总收入滚动环比增长(%)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].operating_total_revenue_lrr_sq', '-', ''),
    (22, 'profit_deduct_nrgal_lrr_sq', '扣非净利润滚动环比增长(%)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].profit_deduct_nrgal_lrr_sq', '-', ''),
    (23, 'wgt_avg_roe', '加权净资产收益率(%)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].wgt_avg_roe', '-', ''),
    (24, 'net_profit_per_share', '每股净资产(元)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].net_profit_per_share', '-', ''),
    (25, 'ncf_from_oa_to_revenue', '经营现金流/营业收入', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].ncf_from_oa_to_revenue', '-', ''),
    (26, 'profit_nrgal_sq', '扣非净利润(元)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].profit_nrgal_sq', '-', ''),
    (27, 'basic_eps', '基本每股收益(元)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].basic_eps', '-', ''),
    (28, 'net_selling_rate', '净利率(%)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].net_selling_rate', '-', ''),
    (29, 'total_capital_turnover', '总资产周转率(次)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].total_capital_turnover', '-', ''),
    (30, 'net_profit_atsopc_lrr_sq', '归属净利润滚动环比增长(%)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].net_profit_atsopc_lrr_sq', '-', ''),
    (31, 'inventory_turnover_days', '存货周转天数(天)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].inventory_turnover_days', '-', ''),
    (32, 'pre_receivable', '预收款/营业收入', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].pre_receivable', '-', ''),
    (33, 'total_revenue', '营业总收入(元)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].total_revenue', '-', '0是有效值不转NULL'),
    (34, 'undistri_profit_ps', '每股未分配利润(元)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].undistri_profit_ps', '-', ''),
    (35, 'dlt_earnings_per_share', '稀释每股收益(元)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].dlt_earnings_per_share', '-', ''),
    (36, 'net_profit_atsopc', '归属净利润(元)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].net_profit_atsopc', '-', ''),
    (37, 'basic_e_ps_net_of_nrgal', '扣非每股收益(元)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].basic_e_ps_net_of_nrgal', '-', ''),
    (38, 'capital_reserve', '每股公积金(元)', 'DECIMAL', '24,4', 'N', 'Y', '', '天眼查', 'result[].capital_reserve', '-', ''),
]

write_sheet(ws9, overview9, fields9)

# ========== Sheet10: company_1114_lawsuit_info ==========
ws10 = wb.create_sheet('company_1114_lawsuit_info')

overview10 = [
    ('数据库', 'powerlink'), ('表名', 'company_1114_lawsuit_info'), ('表描述', '法律诉讼(1114接口)'),
    ('引擎', 'InnoDB'), ('字符集', 'utf8mb4'), ('所属系统', '天眼查数据接入'),
    ('数据关系', '1:N(1公司→N条诉讼)'), ('创建日期', '2026-05-15'),
    ('解析规则说明', 'result.items数组展平,每条诉讼一行; casePersons取前2人展开6列; DELETE旧数据+INSERT新数据; step1含翻页合并存储(天眼查最多500条)'),
    ('特别备注', '1114接口支持翻页(pageNum/pageSize), step1循环翻页合并存一条api_call_record(JSON类型), 保守方案'),
]

fields10 = [
    (1, 'id', '主键ID', 'BIGINT', '', 'Y', 'N', '自增', '内部', '-', '-', '自增主键'),
    (2, 'api_record_id', 'API调用记录ID', 'BIGINT', '', 'N', 'Y', '', '内部', 'api_call_record.id', '-', '关联api_call_record表'),
    (3, 'data_create_time', '数据创建时间', 'DATETIME', '', 'N', 'Y', 'CURRENT_TIMESTAMP', '内部', '-', '-', '解析入库时间'),
    (4, 'company_name', '主公司名(搜索关键字)', 'VARCHAR', '200', 'N', 'N', '', '内部', '-', '-', '来自搜索入参(input_param)，非API返回'),
    (5, 'total', '诉讼记录总数', 'INT', '', 'N', 'Y', '', '天眼查', 'result.total', 'meta字段', '天眼查最多返回500条'),
    (6, 'lawsuit_id', '诉讼条目ID', 'BIGINT', '', 'N', 'Y', '', '天眼查', 'result.items[].id', '字段映射: id→lawsuit_id', '避免与表主键冲突'),
    (7, 'doc_type', '文书类型', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.items[].docType', '驼峰→下划线; [枚举:ENUM_DOC_TYPE(推断)]', '如:判决书/裁定书/调解书等'),
    (8, 'lawsuit_url', '天眼查URL-Web', 'VARCHAR', '500', 'N', 'Y', '', '天眼查', 'result.items[].lawsuitUrl', '驼峰→下划线', ''),
    (9, 'lawsuit_h5_url', '天眼查URL-H5', 'VARCHAR', '500', 'N', 'Y', '', '天眼查', 'result.items[].lawsuitH5Url', '驼峰→下划线', ''),
    (10, 'title', '案件名称', 'VARCHAR', '1000', 'N', 'Y', '', '天眼查', 'result.items[].title', '-', ''),
    (11, 'court', '审理法院', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.items[].court', '-', ''),
    (12, 'judge_time', '裁判日期', 'VARCHAR', '50', 'N', 'Y', '', '天眼查', 'result.items[].judgeTime', '驼峰→下划线', '日期字符串'),
    (13, 'uuid', 'UUID', 'VARCHAR', '100', 'N', 'Y', '', '天眼查', 'result.items[].uuid', '-', ''),
    (14, 'case_no', '案号', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.items[].caseNo', '驼峰→下划线', '如: (2021)京0108民初44534号'),
    (15, 'case_type', '案件类型', 'VARCHAR', '100', 'N', 'Y', '', '天眼查', 'result.items[].caseType', '驼峰→下划线; [枚举:ENUM_CASE_TYPE(推断)]', '如:民事案件/执行案件等'),
    (16, 'case_reason', '案由', 'VARCHAR', '500', 'N', 'Y', '', '天眼查', 'result.items[].caseReason', '驼峰→下划线', '如: 网络侵权责任纠纷'),
    (17, 'case_money', '案件金额', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.items[].caseMoney', '驼峰→下划线; 空字符串→NULL', ''),
    (18, 'submit_time', '发布日期', 'DATETIME', '', 'N', 'Y', '', '天眼查', 'result.items[].submitTime', '毫秒时间戳→datetime(≥1e10÷1000)', ''),
    (19, 'case_result', '案件结果标签', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.items[].casePersons[0].result', '字段映射; [枚举:ENUM_CASE_RESULT(推断)]', '如:胜诉/败诉/撤诉等'),
    (20, 'role1', '案件身份1', 'VARCHAR', '100', 'N', 'Y', '', '天眼查', 'result.items[].casePersons[0].role', 'casePersons展开; [枚举:ENUM_CASE_ROLE(推断)]', '如:原告/被告/上诉人等'),
    (21, 'gid1', 'ID1', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.items[].casePersons[0].gid', 'casePersons展开', '空字符串→NULL'),
    (22, 'emotion1', '情感倾向1', 'INT', '', 'N', 'Y', '', '天眼查', 'result.items[].casePersons[0].emotion', 'casePersons展开; [枚举:ENUM_EMOTION]', '1=正面,0=中性,-1=负面'),
    (23, 'sptname1', '疑似名称1', 'VARCHAR', '500', 'N', 'Y', '', '天眼查', 'result.items[].casePersons[0].sptname', 'casePersons展开; 空字符串→NULL', ''),
    (24, 'name1', '名称1', 'VARCHAR', '500', 'N', 'Y', '', '天眼查', 'result.items[].casePersons[0].name', 'casePersons展开', ''),
    (25, 'type1', '类型1', 'VARCHAR', '50', 'N', 'Y', '', '天眼查', 'result.items[].casePersons[0].type', 'casePersons展开; [枚举:ENUM_SUBJECT_TYPE]', '1=公司,2=人'),
    (26, 'role2', '案件身份2', 'VARCHAR', '100', 'N', 'Y', '', '天眼查', 'result.items[].casePersons[1].role', 'casePersons展开; [枚举:ENUM_CASE_ROLE(推断)]', '如:被告/被上诉人等'),
    (27, 'gid2', 'ID2', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.items[].casePersons[1].gid', 'casePersons展开', '空字符串→NULL'),
    (28, 'emotion2', '情感倾向2', 'INT', '', 'N', 'Y', '', '天眼查', 'result.items[].casePersons[1].emotion', 'casePersons展开; [枚举:ENUM_EMOTION]', '1=正面,0=中性,-1=负面'),
    (29, 'sptname2', '疑似名称2', 'VARCHAR', '500', 'N', 'Y', '', '天眼查', 'result.items[].casePersons[1].sptname', 'casePersons展开; 空字符串→NULL', ''),
    (30, 'name2', '名称2', 'VARCHAR', '500', 'N', 'Y', '', '天眼查', 'result.items[].casePersons[1].name', 'casePersons展开', ''),
    (31, 'type2', '类型2', 'VARCHAR', '50', 'N', 'Y', '', '天眼查', 'result.items[].casePersons[1].type', 'casePersons展开; [枚举:ENUM_SUBJECT_TYPE]', '1=人员, 2=公司'),
]

write_sheet(ws10, overview10, fields10)

# ========== Sheet11: company_973_cash_flow_info ==========
ws11 = wb.create_sheet('company_973_cash_flow_info')

overview11 = [
    ('数据库', 'powerlink'), ('表名', 'company_973_cash_flow_info'), ('表描述', '现金流量表(973接口)'),
    ('引擎', 'InnoDB'), ('字符集', 'utf8mb4'), ('所属系统', '天眼查数据接入'),
    ('数据关系', '1:N(1公司→N个报告期)'), ('创建日期', '2026-05-15'),
    ('解析规则说明', 'result.corpCashFlow数组展平,每报告期一行; 37个VARCHAR字段(带单位如"7.92亿"); 不提取corpFinancialYears; API字段名已是snake_case与DB一致; 空字符串→NULL; DELETE旧数据+INSERT新数据'),
    ('特别备注', '973接口默认返回最近一期数据(如"2026(Q1)"),无需翻页; 非上市公司返回error_code=300000→step2天然跳过'),
]

fields11 = [
    (1, 'id', '主键ID', 'BIGINT', '', 'Y', 'N', '自增', '内部', '-', '-', '自增主键'),
    (2, 'api_record_id', 'API调用记录ID', 'BIGINT', '', 'N', 'Y', '', '内部', 'api_call_record.id', '-', '关联api_call_record表'),
    (3, 'data_create_time', '数据创建时间', 'DATETIME', '', 'N', 'Y', 'CURRENT_TIMESTAMP', '内部', '-', '-', '解析入库时间'),
    (4, 'company_name', '主公司名(搜索关键字)', 'VARCHAR', '200', 'N', 'N', '', '内部', '-', '-', '来自搜索入参(input_param)，非API返回'),
    (5, 'show_year', '报告期', 'VARCHAR', '32', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].showYear', '字段映射: showYear→show_year', '如: 2026(Q1), 2025(年报)'),
    (6, 'ncf_from_oa', '经营活动产生的现金流量净额', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].ncf_from_oa', 'API已是snake_case', ''),
    (7, 'sub_total_of_ci_from_oa', '经营活动现金流入小计', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].sub_total_of_ci_from_oa', 'API已是snake_case', ''),
    (8, 'sub_total_of_cos_from_oa', '经营活动现金流出小计', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].sub_total_of_cos_from_oa', 'API已是snake_case', ''),
    (9, 'cash_received_of_sales_service', '销售商品、提供劳务收到的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].cash_received_of_sales_service', 'API已是snake_case', ''),
    (10, 'payments_of_all_taxes', '支付的各项税费', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].payments_of_all_taxes', 'API已是snake_case', ''),
    (11, 'cash_paid_to_staff_etc', '支付给职工以及为职工支付的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].cash_paid_to_staff_etc', 'API已是snake_case', ''),
    (12, 'goods_buy_and_service_cash_pay', '购买商品、接受劳务支付的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].goods_buy_and_service_cash_pay', 'API已是snake_case', ''),
    (13, 'other_cash_paid_related_to_oa', '支付其他与经营活动有关的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].other_cash_paid_related_to_oa', 'API已是snake_case; 空字符串→NULL', ''),
    (14, 'cash_received_of_other_fa', '收到其他与经营活动有关的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].cash_received_of_other_fa', 'API已是snake_case(虽名含fa实为经营活动)', 'API命名遗留，实为经营活动'),
    (15, 'ncf_from_ia', '投资活动产生的现金流量净额', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].ncf_from_ia', 'API已是snake_case', ''),
    (16, 'sub_total_of_ci_from_ia', '投资活动现金流入小计', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].sub_total_of_ci_from_ia', 'API已是snake_case', ''),
    (17, 'sub_total_of_cos_from_ia', '投资活动现金流出小计', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].sub_total_of_cos_from_ia', 'API已是snake_case', ''),
    (18, 'cash_received_of_dspsl_invest', '收回投资收到的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].cash_received_of_dspsl_invest', 'API已是snake_case(Dspsl=Disposal缩写)', ''),
    (19, 'invest_income_cash_received', '取得投资收益收到的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].invest_income_cash_received', 'API已是snake_case', ''),
    (20, 'net_cash_of_disposal_assets', '处置固定资产等收回的现金净额', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].net_cash_of_disposal_assets', 'API已是snake_case', ''),
    (21, 'net_cash_of_disposal_branch', '处置子公司等收到的现金净额', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].net_cash_of_disposal_branch', 'API已是snake_case; 空字符串→NULL', ''),
    (22, 'cash_received_of_other_ia', '收到其他与投资活动有关的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].cash_received_of_other_ia', 'API已是snake_case', ''),
    (23, 'invest_paid_cash', '投资支付的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].invest_paid_cash', 'API已是snake_case', ''),
    (24, 'cash_paid_for_assets', '购建固定资产等支付的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].cash_paid_for_assets', 'API已是snake_case', ''),
    (25, 'ncf_from_fa', '筹资活动产生的现金流量净额', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].ncf_from_fa', 'API已是snake_case', ''),
    (26, 'sub_total_of_ci_from_fa', '筹资活动现金流入小计', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].sub_total_of_ci_from_fa', 'API已是snake_case', ''),
    (27, 'sub_total_of_cos_from_fa', '筹资活动现金流出小计', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].sub_total_of_cos_from_fa', 'API已是snake_case', ''),
    (28, 'cash_received_of_absorb_invest', '吸收投资收到的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].cash_received_of_absorb_invest', 'API已是snake_case', ''),
    (29, 'cash_received_from_investor', '子公司吸收少数股东投资收到的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].cash_received_from_investor', 'API已是snake_case', ''),
    (30, 'cash_received_of_borrowing', '取得借款收到的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].cash_received_of_borrowing', 'API已是snake_case', ''),
    (31, 'cash_received_from_bond_issue', '发行债券收到的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].cash_received_from_bond_issue', 'API已是snake_case; 空字符串→NULL', ''),
    (32, 'cash_received_of_othr_fa', '收到其他与筹资活动有关的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].cash_received_of_othr_fa', 'API已是snake_case(Othr=Other缩写)', ''),
    (33, 'cash_pay_for_debt', '偿还债务支付的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].cash_pay_for_debt', 'API已是snake_case', ''),
    (34, 'cash_paid_of_distribution', '分配股利、利润或偿付利息支付的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].cash_paid_of_distribution', 'API已是snake_case', ''),
    (35, 'other_cash_paid_relating_to_fa', '支付其他与筹资活动有关的现金', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].other_cash_paid_relating_to_fa', 'API已是snake_case', ''),
    (36, 'branch_paid_to_minority_holder', '子公司支付给少数股东的股利、利润', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].branch_paid_to_minority_holder', 'API已是snake_case; 空字符串→NULL', ''),
    (37, 'net_increase_in_cce', '现金及现金等价物净增加额', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].net_increase_in_cce', 'API已是snake_case(CCE=CashEquivalents)', ''),
    (38, 'initial_balance_of_cce', '期初现金及现金等价物余额', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].initial_balance_of_cce', 'API已是snake_case', ''),
    (39, 'final_balance_of_cce', '期末现金及现金等价物余额', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].final_balance_of_cce', 'API已是snake_case', ''),
    (40, 'net_cash_amt_from_branch', '取得子公司等支付的现金净额', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].net_cash_amt_from_branch', 'API已是snake_case', ''),
    (41, 'effect_of_exchange_chg_on_cce', '汇率变动对现金等价物的影响', 'VARCHAR', '200', 'N', 'Y', '', '天眼查', 'result.corpCashFlow[].effect_of_exchange_chg_on_cce', 'API已是snake_case', ''),
]

write_sheet(ws11, overview11, fields11)

# ========== Sheet12: company_P51060_paydex_info ==========
ws12 = wb.create_sheet('company_P51060_paydex_info')

overview12 = [
    ('数据库', 'powerlink'), ('表名', 'company_P51060_paydex_info'), ('表描述', '付款指数(邓白氏P51060接口)'),
    ('引擎', 'InnoDB'), ('字符集', 'utf8mb4'), ('所属系统', '邓白氏数据接入'),
    ('唯一约束', 'company_name'), ('创建日期', '2026-05-18'),
    ('解析规则说明', '1:1关系, ON DUPLICATE KEY UPDATE; 邓白氏res为JSON字符串需二次解析; companyHistoryPayDexes(List)→JSON字符串存储; company_name来自搜索入参(entityName); 空字符串→NULL'),
    ('与天眼查差异', 'POST请求+SHA256签名认证; 搜索参数entityName+uscc(从819表查取); 响应格式{code,res,msg,trace}中res为JSON字符串; 成功判断code=0'),
]

fields12 = [
    (1, 'id', '主键ID', 'BIGINT', '', 'Y', 'N', '自增', '内部', '-', '-', '自增主键'),
    (2, 'api_record_id', 'API调用记录ID', 'BIGINT', '', 'N', 'Y', '', '内部', 'api_call_record.id', '-', '关联api_call_record表'),
    (3, 'data_create_time', '数据创建时间', 'DATETIME', '', 'N', 'Y', 'CURRENT_TIMESTAMP', '内部', '-', '-', '解析入库时间'),
    (4, 'company_name', '主公司名(搜索关键字)', 'VARCHAR', '200', 'UK', 'N', '', '内部', '-', '-', '来自搜索入参(entityName)，非API返回'),
    (5, 'uscc', '统一社会信用代码', 'VARCHAR', '50', 'N', 'Y', '', '邓白氏', 'res.uscc', '空字符串→NULL', 'API返回的统一社会信用代码'),
    (6, 'company_paydex', 'PayDex评分数值(最新)', 'VARCHAR', '50', 'N', 'Y', '', '邓白氏', 'res.companyPayDex', '驼峰→下划线; [枚举:ENUM_PAYDEX_SCORE(推断)]', '0-100评分,80=按时,50-69=逾期'),
    (7, 'company_paydex_date', 'PayDex评分日期(最新)', 'VARCHAR', '20', 'N', 'Y', '', '邓白氏', 'res.companyPayDexDate', '驼峰→下划线; 空字符串→NULL', '如: 2025-0509'),
    (8, 'company_history_paydexes', 'PayDex历史信息', 'TEXT', '', 'N', 'Y', '', '邓白氏', 'res.companyHistoryPayDexes', 'List→JSON字符串存储', 'JSON数组，含历史评分变化'),
    (9, 'sic2', 'SIC前2位', 'VARCHAR', '20', 'N', 'Y', '', '邓白氏', 'res.sic2', '空字符串→NULL', '标准行业分类代码前2位'),
    (10, 'sic3', 'SIC前3位', 'VARCHAR', '20', 'N', 'Y', '', '邓白氏', 'res.sic3', '空字符串→NULL', '标准行业分类代码前3位'),
    (11, 'sic4', 'SIC前4位', 'VARCHAR', '20', 'N', 'Y', '', '邓白氏', 'res.sic4', '空字符串→NULL', '标准行业分类代码前4位'),
    (12, 'industry_paydex_date', '行业PayDex评分日期(最新)', 'VARCHAR', '20', 'N', 'Y', '', '邓白氏', 'res.industryPayDexDate', '驼峰→下划线; 空字符串→NULL', '行业评分更新日期'),
    (13, 'industry_lower_quartile_paydex', '行业25分位PayDex评分', 'VARCHAR', '50', 'N', 'Y', '', '邓白氏', 'res.industryLowerQuartilePayDex', '驼峰→下划线; 空字符串→NULL', '行业下四分位值'),
    (14, 'industry_median_paydex', '行业50分位PayDex评分', 'VARCHAR', '50', 'N', 'Y', '', '邓白氏', 'res.industryMedianPayDex', '驼峰→下划线; 空字符串→NULL', '行业中位值'),
    (15, 'industry_upper_quartile_paydex', '行业75分位PayDex评分', 'VARCHAR', '50', 'N', 'Y', '', '邓白氏', 'res.industryUpperQuartilePayDex', '驼峰→下划线; 空字符串→NULL', '行业上四分位值'),
    (16, 'industry_count_num', '行业统计数据-样本数量', 'VARCHAR', '50', 'N', 'Y', '', '邓白氏', 'res.industryCountNum', '驼峰→下划线; 空字符串→NULL', '行业统计的企业主体样本数量'),
    (17, 'industry_company_position', '行业位置', 'VARCHAR', '50', 'N', 'Y', '', '邓白氏', 'res.industryCompanyPosition', '驼峰→下划线; 空字符串→NULL', '企业在行业中的百分位位置'),
    (18, 'company_average', '平均付款天数(中文)', 'VARCHAR', '100', 'N', 'Y', '', '邓白氏', 'res.companyAverage', '驼峰→下划线; 空字符串→NULL', '如: 逾期30天'),
    (19, 'en_company_average', '平均付款天数(英文)', 'VARCHAR', '100', 'N', 'Y', '', '邓白氏', 'res.encompanyAverage', '驼峰→下划线; en前缀保留; 空字符串→NULL', '如: 30 days beyond terms'),
    (20, 'industry_average', '行业平均付款天数(中文)', 'VARCHAR', '100', 'N', 'Y', '', '邓白氏', 'res.industryAverage', '驼峰→下划线; 空字符串→NULL', '如: 逾期19天'),
    (21, 'en_industry_average', '行业平均付款天数(英文)', 'VARCHAR', '100', 'N', 'Y', '', '邓白氏', 'res.enindustryAverage', '驼峰→下划线; en前缀保留; 空字符串→NULL', '如: 19 days beyond terms'),
]

write_sheet(ws12, overview12, fields12)

# ========== 重排Sheet顺序 ==========
# 规则：非接口表 → 接口记录表 → 接口解析表(按接口号升序)
desired_order = [
    'customer_info',           # 非接口表(数据源)
    'api_call_record',         # 接口记录表
    'company_819_info',        # 819
    'company_822_change_info', # 822
    'company_854_stock_info',  # 854
    'company_1058_risk_info',  # 1058
    'company_1114_lawsuit_info', # 1114
    'company_1149_scale_info', # 1149
    'company_1168_org_type_info', # 1168
    'company_967_main_index_info', # 967
    'company_973_cash_flow_info',  # 973
    'company_P51060_paydex_info',  # P51060(邓白氏)
    '枚举字典',                  # 枚举字典
]
# 按desired_order重排wb._sheets
sheet_map = {ws.title: ws for ws in wb.worksheets}
wb._sheets = [sheet_map[name] for name in desired_order if name in sheet_map]

# ========== 目录Sheet（插入到最前面） ==========

# 目录数据：序号、表名、中文名、字段数、数据关系、接口号、说明
# 按重排后的顺序：非接口表 → 接口记录表 → 接口解析表(按接口号升序)
dir_data = [
    ('customer_info', '客户公司列表', 3, '-', '-', '数据源，提供搜索关键字'),
    ('api_call_record', '三方接口调用记录', 7, '1:N(所有接口共用)', '-', '所有接口的API调用原始记录，通过interface_name区分'),
    ('company_819_info', '企业基本信息', 65, '1:1', '819', '含主要人员、行业分类、资本、证件编码等'),
    ('company_822_change_info', '变更记录', 10, '1:N', '822', '2层展平：变更总数+每条变更记录'),
    ('company_854_stock_info', '上市公司企业简介', 36, '1:1', '854', '4个Object人物字段展开+非上市公司天然跳过'),
    ('company_1058_risk_info', '企业天眼风险', 16, '1:N', '1058', '3层嵌套展平：风险类别→风险类型→风险条目'),
    ('company_1114_lawsuit_info', '法律诉讼', 31, '1:N', '1114', '翻页合并+casePersons前2人展开'),
    ('company_1149_scale_info', '企业规模', 5, '1:1', '1149', 'result直接为字符串(如"大型")'),
    ('company_1168_org_type_info', '组织机构类型', 7, '1:1', '1168', 'orgTypes/economyTypes数组→逗号分隔拆列'),
    ('company_967_main_index_info', '主要指标-年度', 38, '1:N', '967', '~28个DECIMAL字段，每年度一行'),
    ('company_973_cash_flow_info', '现金流量表', 41, '1:N', '973', '37个VARCHAR字段+showYear，API默认返回最近一期'),
    ('company_P51060_paydex_info', '付款指数', 21, '1:1', 'P51060', '邓白氏PAYDEX评分+行业基准+历史评分(JSON); POST+SHA256签名认证'),
    ('枚举字典', '枚举值定义', '-', '-', '-', '所有接口的枚举/固定值字段定义汇总'),
]

ws_dir = wb.create_sheet('目录', 0)  # 插入到位置0（最前面）

# 标题
ws_dir.merge_cells('A1:G1')
ws_dir['A1'] = 'PowerLink 数据字典目录'
ws_dir['A1'].font = Font(name='微软雅黑', bold=True, size=16, color='1F4E79')
ws_dir['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_dir['A1'].fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
ws_dir['A1'].border = thin_border
ws_dir.row_dimensions[1].height = 36

# 表头
dir_headers = ['序号', '表名(点击跳转)', '中文描述', '字段数', '数据关系', '接口号', '说明']
dir_col_widths = [6, 30, 16, 8, 18, 8, 40]

header_row = 3
for col_idx, h in enumerate(dir_headers, 1):
    cell = ws_dir.cell(row=header_row, column=col_idx)
    cell.value = h
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# 数据行（带超链接）
for row_idx, (table_name, chinese, field_count, relation, api_num, desc) in enumerate(dir_data):
    r = header_row + 1 + row_idx
    ws_dir.cell(row=r, column=1, value=row_idx + 1).font = normal_font
    ws_dir.cell(row=r, column=1).alignment = center_align
    ws_dir.cell(row=r, column=1).border = thin_border

    # 表名+超链接（使用location格式，内部链接不带#前缀）
    link_cell = ws_dir.cell(row=r, column=2, value=table_name)
    link_cell.font = Font(name='微软雅黑', size=10, color='2E75B6', underline='single')
    link_cell.alignment = left_align
    link_cell.border = thin_border
    link_cell.hyperlink = Hyperlink(ref=f'B{r}', location=f'{table_name}!A1', display=table_name)

    ws_dir.cell(row=r, column=3, value=chinese).font = normal_font
    ws_dir.cell(row=r, column=3).alignment = center_align
    ws_dir.cell(row=r, column=3).border = thin_border

    ws_dir.cell(row=r, column=4, value=field_count).font = normal_font
    ws_dir.cell(row=r, column=4).alignment = center_align
    ws_dir.cell(row=r, column=4).border = thin_border

    ws_dir.cell(row=r, column=5, value=relation).font = normal_font
    ws_dir.cell(row=r, column=5).alignment = center_align
    ws_dir.cell(row=r, column=5).border = thin_border

    ws_dir.cell(row=r, column=6, value=api_num).font = normal_font
    ws_dir.cell(row=r, column=6).alignment = center_align
    ws_dir.cell(row=r, column=6).border = thin_border

    ws_dir.cell(row=r, column=7, value=desc).font = normal_font
    ws_dir.cell(row=r, column=7).alignment = left_align
    ws_dir.cell(row=r, column=7).border = thin_border

    if row_idx % 2 == 1:
        for col in range(1, 8):
            ws_dir.cell(row=r, column=col).fill = alt_fill

# 列宽
for i, w in enumerate(dir_col_widths, 1):
    ws_dir.column_dimensions[get_column_letter(i)].width = w


# ========== Sheet13: 枚举字典 ==========
ws_enum = wb.create_sheet('枚举字典')

overview_enum = [
    ('数据库', 'powerlink'), ('表名', '-(虚拟表)'), ('表描述', '所有接口的枚举值/固定值定义汇总'),
    ('数据来源', '天眼查9个接口+邓白氏1个接口的API文档PDF'), ('确定性说明', '明确=文档中有完整值定义; 推断=文档仅有示例或业务常识推断,非完整枚举'),
]

enum_headers = ['序号', '枚举名称', '适用接口', '适用DB列', '枚举值', '值含义', '确定性', '来源']
enum_col_widths = [6, 22, 14, 35, 10, 25, 8, 40]

# 枚举数据：每行(序号, 枚举名称, 适用接口, 适用DB列, 枚举值, 值含义, 确定性, 来源)
enum_fields = [
    # --- ENUM_LEGAL_PERSON_TYPE (819 法人类型) ---
    (1, 'ENUM_LEGAL_PERSON_TYPE', '819', 'legal_person_type', '1', '自然人', '明确', '819接口文档'),
    (2, 'ENUM_LEGAL_PERSON_TYPE', '819', 'legal_person_type', '2', '公司', '明确', '819接口文档'),
    # --- ENUM_SUBJECT_TYPE (854/1114 主体类型) ---
    (3, 'ENUM_SUBJECT_TYPE', '854/1114', 'gm_type,chairman_type,secretary_type,legal_person_type(854); type1,type2(1114)', '1', '公司', '明确', '854/1114接口文档'),
    (4, 'ENUM_SUBJECT_TYPE', '854/1114', '', '2', '人', '明确', '854/1114接口文档'),
    # --- ENUM_IS_MICRO_ENT (819 是否小微企业) ---
    (5, 'ENUM_IS_MICRO_ENT', '819', 'is_micro_ent', '0', '否', '明确', '819接口文档'),
    (6, 'ENUM_IS_MICRO_ENT', '819', 'is_micro_ent', '1', '是', '明确', '819接口文档'),
    # --- ENUM_RISK_CATEGORY (1058 风险分类) ---
    (7, 'ENUM_RISK_CATEGORY', '1058', 'risk_category_name(间接)', '0', '预警提醒', '明确', '1058接口文档'),
    (8, 'ENUM_RISK_CATEGORY', '1058', '', '1', '自身风险', '明确', '1058接口文档'),
    (9, 'ENUM_RISK_CATEGORY', '1058', '', '2', '周边风险', '明确', '1058接口文档'),
    (10, 'ENUM_RISK_CATEGORY', '1058', '', '3', '历史风险', '明确', '1058接口文档'),
    # --- ENUM_RISK_TAG (1058 风险标签) ---
    (11, 'ENUM_RISK_TAG', '1058', 'risk_type_tag', '高风险', '高风险', '明确', '1058接口文档'),
    (12, 'ENUM_RISK_TAG', '1058', 'risk_type_tag', '警示', '警示', '明确', '1058接口文档'),
    (13, 'ENUM_RISK_TAG', '1058', 'risk_type_tag', '提示信息', '提示信息', '明确', '1058接口文档'),
        # --- ENUM_RISK_TYPE (1058 风险类型码,完整72个值) ---
    (14, 'ENUM_RISK_TYPE', '1058', 'risk_type', '1', '严重违法_企业', '明确', '1058接口文档'),
    (15, 'ENUM_RISK_TYPE', '1058', '', '3', '失信被执行人_企业', '明确', ''),
    (16, 'ENUM_RISK_TYPE', '1058', '', '5', '被执行人_企业', '明确', ''),
    (17, 'ENUM_RISK_TYPE', '1058', '', '6', '行政处罚_企业', '明确', ''),
    (18, 'ENUM_RISK_TYPE', '1058', '', '7', '经营异常_企业', '明确', ''),
    (19, 'ENUM_RISK_TYPE', '1058', '', '8', '裁判文书_企业', '明确', ''),
    (20, 'ENUM_RISK_TYPE', '1058', '', '9', '股权出质_企业', '明确', ''),
    (21, 'ENUM_RISK_TYPE', '1058', '', '10', '动产抵押_企业', '明确', ''),
    (22, 'ENUM_RISK_TYPE', '1058', '', '11', '欠税公告_企业', '明确', ''),
    (23, 'ENUM_RISK_TYPE', '1058', '', '12', '名称变更_企业', '明确', ''),
    (24, 'ENUM_RISK_TYPE', '1058', '', '13', '开庭公告_企业', '明确', ''),
    (25, 'ENUM_RISK_TYPE', '1058', '', '14', '法院公告_企业', '明确', ''),
    (26, 'ENUM_RISK_TYPE', '1058', '', '15', '法人变更_企业', '明确', ''),
    (27, 'ENUM_RISK_TYPE', '1058', '', '16', '投资人变更_企业', '明确', ''),
    (28, 'ENUM_RISK_TYPE', '1058', '', '17', '主要人员变更_企业', '明确', ''),
    (29, 'ENUM_RISK_TYPE', '1058', '', '18', '注册资本变更_企业', '明确', ''),
    (30, 'ENUM_RISK_TYPE', '1058', '', '19', '注册地址变更_企业', '明确', ''),
    (31, 'ENUM_RISK_TYPE', '1058', '', '20', '出资情况变更_企业', '明确', ''),
    (32, 'ENUM_RISK_TYPE', '1058', '', '21', '股权冻结_企业', '明确', ''),
    (33, 'ENUM_RISK_TYPE', '1058', '', '22', '清算信息_企业', '明确', ''),
    (34, 'ENUM_RISK_TYPE', '1058', '', '23', '知识产权出质_企业', '明确', ''),
    (35, 'ENUM_RISK_TYPE', '1058', '', '24', '环保处罚_企业', '明确', ''),
    (36, 'ENUM_RISK_TYPE', '1058', '', '25', '公示催告_企业', '明确', ''),
    (37, 'ENUM_RISK_TYPE', '1058', '', '26', '送达公告_企业', '明确', ''),
    (38, 'ENUM_RISK_TYPE', '1058', '', '27', '立案信息_企业', '明确', ''),
    (39, 'ENUM_RISK_TYPE', '1058', '', '28', '税收违法_企业', '明确', ''),
    (40, 'ENUM_RISK_TYPE', '1058', '', '29', '司法拍卖_企业', '明确', ''),
    (41, 'ENUM_RISK_TYPE', '1058', '', '30', '土地抵押_企业', '明确', ''),
    (42, 'ENUM_RISK_TYPE', '1058', '', '31', '简易注销_企业', '明确', ''),
    (43, 'ENUM_RISK_TYPE', '1058', '', '32', '限制消费令_企业', '明确', ''),
    (44, 'ENUM_RISK_TYPE', '1058', '', '33', '限制消费令_人员', '明确', ''),
    (45, 'ENUM_RISK_TYPE', '1058', '', '34', '终本案件_企业', '明确', ''),
    (46, 'ENUM_RISK_TYPE', '1058', '', '35', '股权出质_人员', '明确', ''),
    (47, 'ENUM_RISK_TYPE', '1058', '', '36', '股权冻结_人员', '明确', ''),
    (48, 'ENUM_RISK_TYPE', '1058', '', '37', '股权质押_人员', '明确', ''),
    (49, 'ENUM_RISK_TYPE', '1058', '', '38', '破产案件_企业', '明确', ''),
    (50, 'ENUM_RISK_TYPE', '1058', '', '39', '询价评估_企业', '明确', ''),
    (51, 'ENUM_RISK_TYPE', '1058', '', '40', '抽查检查_企业', '明确', ''),
    (52, 'ENUM_RISK_TYPE', '1058', '', '41', '对外担保_企业', '明确', ''),
    (53, 'ENUM_RISK_TYPE', '1058', '', '42', '违规处理_企业', '明确', ''),
    (54, 'ENUM_RISK_TYPE', '1058', '', '45', '强制清算_企业', '明确', ''),
    (55, 'ENUM_RISK_TYPE', '1058', '', '46', '终本案件_人员', '明确', ''),
    (56, 'ENUM_RISK_TYPE', '1058', '', '47', '开庭公告_人员', '明确', ''),
    (57, 'ENUM_RISK_TYPE', '1058', '', '48', '法院公告_人员', '明确', ''),
    (58, 'ENUM_RISK_TYPE', '1058', '', '49', '送达公告_人员', '明确', ''),
    (59, 'ENUM_RISK_TYPE', '1058', '', '50', '立案信息_人员', '明确', ''),
    (60, 'ENUM_RISK_TYPE', '1058', '', '51', '股权质押_企业', '明确', ''),
    (61, 'ENUM_RISK_TYPE', '1058', '', '53', '严重违法(已移出)_企业', '明确', ''),
    (62, 'ENUM_RISK_TYPE', '1058', '', '55', '经营异常(已移出)_企业', '明确', ''),
    (63, 'ENUM_RISK_TYPE', '1058', '', '56', '裁判文书_人员', '明确', ''),
    (64, 'ENUM_RISK_TYPE', '1058', '', '63', '注销备案_企业', '明确', ''),
    (65, 'ENUM_RISK_TYPE', '1058', '', '64', '食品安全_企业', '明确', ''),
    (66, 'ENUM_RISK_TYPE', '1058', '', '65', '产品召回_企业', '明确', ''),
    (67, 'ENUM_RISK_TYPE', '1058', '', '70', '行政处罚_企业_历史', '明确', ''),
    (68, 'ENUM_RISK_TYPE', '1058', '', '71', '失信被执行人_企业_历史', '明确', ''),
    (69, 'ENUM_RISK_TYPE', '1058', '', '72', '被执行人_企业_历史', '明确', ''),
    (70, 'ENUM_RISK_TYPE', '1058', '', '73', '限制消费令_企业_历史', '明确', ''),
    (71, 'ENUM_RISK_TYPE', '1058', '', '74', '终本案件_企业_历史', '明确', ''),
    (72, 'ENUM_RISK_TYPE', '1058', '', '75', '股权冻结_企业_历史', '明确', ''),
    (73, 'ENUM_RISK_TYPE', '1058', '', '76', '经营异常_企业_历史', '明确', ''),
    (74, 'ENUM_RISK_TYPE', '1058', '', '77', '股权出质_企业_历史', '明确', ''),
    (75, 'ENUM_RISK_TYPE', '1058', '', '78', '动产抵押_企业_历史', '明确', ''),
    (76, 'ENUM_RISK_TYPE', '1058', '', '79', '欠税公告_企业_历史', '明确', ''),
    (77, 'ENUM_RISK_TYPE', '1058', '', '80', '终本案件_人员_历史', '明确', ''),
    (78, 'ENUM_RISK_TYPE', '1058', '', '81', '股权冻结_人员_历史', '明确', ''),
    (79, 'ENUM_RISK_TYPE', '1058', '', '82', '股权出质_人员_历史', '明确', ''),
    (80, 'ENUM_RISK_TYPE', '1058', '', '99', '破产案件_企业(历史)', '明确', ''),
    (81, 'ENUM_RISK_TYPE', '1058', '', '100', '减资公告_企业', '明确', ''),
    (82, 'ENUM_RISK_TYPE', '1058', '', '101', '开庭公告_企业_历史', '明确', ''),
    (83, 'ENUM_RISK_TYPE', '1058', '', '102', '裁判文书_企业_历史', '明确', ''),
    (84, 'ENUM_RISK_TYPE', '1058', '', '103', '法院公告_企业_历史', '明确', ''),
    (85, 'ENUM_RISK_TYPE', '1058', '', '104', '立案信息_企业_历史', '明确', ''),
    # --- ENUM_EMOTION (1114 情感倾向) ---
    (44, 'ENUM_EMOTION', '1114', 'emotion1,emotion2', '1', '正面', '明确', '1114接口文档'),
    (45, 'ENUM_EMOTION', '1114', '', '0', '中性', '明确', '1114接口文档'),
    (46, 'ENUM_EMOTION', '1114', '', '-1', '负面', '明确', '1114接口文档'),
    # --- ENUM_DNB_STATUS_CODE (P51060 邓白氏响应码) ---
    (47, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', 'api_call_record.status_code', '0', '有效请求-成功', '明确', 'P51060接口文档附录'),
    (48, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '1', '有效请求-请求无结果', '明确', ''),
    (49, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '1000', '无效请求-请求参数错误', '明确', ''),
    (50, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '1001', '无效请求-认证参数错误', '明确', ''),
    (51, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '1002', '无效请求-签名验证错误', '明确', ''),
    (52, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '1003', '无效请求-客户IP错误(IP白名单)', '明确', ''),
    (53, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '1004', '无效请求-账号不可用', '明确', ''),
    (54, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '1005', '无效请求-账号已过期', '明确', ''),
    (55, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '1006', '无效请求-账户余额不足', '明确', ''),
    (56, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '1007', '无效请求-访问频繁触发限流', '明确', ''),
    (57, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '1008', '无效请求-接口用量已达日上限', '明确', ''),
    (58, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '1009', '无效请求-产品服务不可用', '明确', ''),
    (59, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '1010', '无效请求-产品服务已过期', '明确', ''),
    (60, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '1011', '无效请求-产品服务已使用完毕', '明确', ''),
    (61, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '1012', '无效请求-产品服务未生效', '明确', ''),
    (62, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '1014', '无效请求-业务逻辑问题', '明确', ''),
    (63, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '2001', '无效请求-处理失败', '明确', ''),
    (64, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '2002', '无效请求-系统错误', '明确', ''),
    (65, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '2003', '无效请求-请求超时', '明确', ''),
    (66, 'ENUM_DNB_STATUS_CODE', 'P51060(邓白氏)', '', '2004', '无效请求-配置错误', '明确', ''),
    # --- 推断枚举 ---
    # --- ENUM_REG_STATUS (819 经营状态,推断) ---
    (67, 'ENUM_REG_STATUS', '819', 'reg_status', '存续', '存续(在营)', '推断', '819接口示例+业务常识'),
    (68, 'ENUM_REG_STATUS', '819', '', '注销', '注销', '推断', ''),
    (69, 'ENUM_REG_STATUS', '819', '', '吊销', '吊销', '推断', ''),
    (70, 'ENUM_REG_STATUS', '819', '', '迁出', '迁出', '推断', ''),
    (71, 'ENUM_REG_STATUS', '819', '', '停业', '停业', '推断', ''),
    # --- ENUM_COMPANY_SCALE (1149 企业规模,推断) ---
    (72, 'ENUM_COMPANY_SCALE', '1149', 'company_scale', '大型', '大型企业', '推断', '1149接口示例+业务常识'),
    (73, 'ENUM_COMPANY_SCALE', '1149', '', '中型', '中型企业', '推断', ''),
    (74, 'ENUM_COMPANY_SCALE', '1149', '', '小型', '小型企业', '推断', ''),
    (75, 'ENUM_COMPANY_SCALE', '1149', '', '微型', '微型企业', '推断', ''),
    # --- ENUM_CURRENCY (819 币种,推断) ---
    (76, 'ENUM_CURRENCY', '819', 'reg_capital_currency,paid_capital_currency', '人民币', '人民币', '推断', '819接口文档标注"人民币 美元 欧元 等"'),
    (77, 'ENUM_CURRENCY', '819', '', '美元', '美元', '推断', ''),
    (78, 'ENUM_CURRENCY', '819', '', '欧元', '欧元', '推断', ''),
    # --- ENUM_DOC_TYPE (1114 文书类型,推断) ---
    (79, 'ENUM_DOC_TYPE', '1114', 'doc_type', '判决书', '判决书', '推断', '1114接口示例+业务常识'),
    (80, 'ENUM_DOC_TYPE', '1114', '', '裁定书', '裁定书', '推断', ''),
    (81, 'ENUM_DOC_TYPE', '1114', '', '调解书', '调解书', '推断', ''),
    (82, 'ENUM_DOC_TYPE', '1114', '', '决定书', '决定书', '推断', ''),
    # --- ENUM_CASE_TYPE (1114 案件类型,推断) ---
    (83, 'ENUM_CASE_TYPE', '1114', 'case_type', '民事案件', '民事案件', '推断', '1114接口示例+业务常识'),
    (84, 'ENUM_CASE_TYPE', '1114', '', '执行案件', '执行案件', '推断', ''),
    (85, 'ENUM_CASE_TYPE', '1114', '', '管辖案件', '管辖案件', '推断', ''),
    (86, 'ENUM_CASE_TYPE', '1114', '', '刑事案件', '刑事案件', '推断', ''),
    # --- ENUM_CASE_ROLE (1114 案件身份,推断) ---
    (87, 'ENUM_CASE_ROLE', '1114', 'role1,role2', '原告', '原告', '推断', '1114接口示例+业务常识'),
    (88, 'ENUM_CASE_ROLE', '1114', '', '被告', '被告', '推断', ''),
    (89, 'ENUM_CASE_ROLE', '1114', '', '上诉人', '上诉人', '推断', ''),
    (90, 'ENUM_CASE_ROLE', '1114', '', '被上诉人', '被上诉人', '推断', ''),
    (91, 'ENUM_CASE_ROLE', '1114', '', '申请执行人', '申请执行人', '推断', ''),
    (92, 'ENUM_CASE_ROLE', '1114', '', '被执行人', '被执行人', '推断', ''),
    # --- ENUM_CASE_RESULT (1114 案件结果,推断) ---
    (93, 'ENUM_CASE_RESULT', '1114', 'case_result', '胜诉', '胜诉', '推断', '1114接口示例+业务常识'),
    (94, 'ENUM_CASE_RESULT', '1114', '', '败诉', '败诉', '推断', ''),
    (95, 'ENUM_CASE_RESULT', '1114', '', '撤诉', '撤诉', '推断', ''),
    (96, 'ENUM_CASE_RESULT', '1114', '', '部分支持', '部分支持', '推断', ''),
    (97, 'ENUM_CASE_RESULT', '1114', '', '驳回', '驳回', '推断', ''),
    # --- ENUM_PAYDEX_SCORE (P51060 PAYDEX评分语义,推断) ---
    (98, 'ENUM_PAYDEX_SCORE', 'P51060(邓白氏)', 'company_paydex', '0-49', '严重逾期', '推断', 'PAYDEX业务常识(PDF无明确阈值定义)'),
    (99, 'ENUM_PAYDEX_SCORE', 'P51060(邓白氏)', '', '50-69', '明显逾期', '推断', ''),
    (100, 'ENUM_PAYDEX_SCORE', 'P51060(邓白氏)', '', '70-79', '稍慢', '推断', ''),
    (101, 'ENUM_PAYDEX_SCORE', 'P51060(邓白氏)', '', '80', '按期及时付款', '推断', ''),
]

# --- 枚举字典专属渲染 ---
start_overview_row = 2
ws_enum[f'A{start_overview_row - 1}'] = '← 返回目录'
ws_enum[f'A{start_overview_row - 1}'].font = Font(name='微软雅黑', bold=True, size=10, color='2E75B6', underline='single')
ws_enum[f'A{start_overview_row - 1}'].hyperlink = Hyperlink(ref=f'A{start_overview_row - 1}', location=f'目录!A1', display='← 返回目录')

# 概览区
overview_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
last_enum_col = get_column_letter(len(enum_headers))
ws_enum.merge_cells(f'A{start_overview_row}:{last_enum_col}{start_overview_row}')
ws_enum[f'A{start_overview_row}'] = '枚举字典概览'
ws_enum[f'A{start_overview_row}'].font = title_font
ws_enum[f'A{start_overview_row}'].alignment = left_align
ws_enum[f'A{start_overview_row}'].fill = overview_fill
ws_enum[f'A{start_overview_row}'].border = thin_border

for i, (k, v) in enumerate(overview_enum):
    r = start_overview_row + 1 + i
    key_cell = ws_enum[f'A{r}']
    key_cell.value = k
    key_cell.font = Font(name='微软雅黑', bold=True, size=10, color='1F4E79')
    key_cell.alignment = left_align
    key_cell.border = thin_border
    key_cell.fill = overview_fill
    val_cell = ws_enum[f'B{r}']
    val_cell.value = v
    val_cell.font = normal_font
    val_cell.alignment = left_align
    val_cell.border = thin_border
    ws_enum.merge_cells(f'B{r}:{last_enum_col}{r}')

# 字段区标题
start_fields_row = start_overview_row + 1 + len(overview_enum) + 1
ws_enum.merge_cells(f'A{start_fields_row}:{last_enum_col}{start_fields_row}')
ws_enum[f'A{start_fields_row}'] = '枚举值明细'
ws_enum[f'A{start_fields_row}'].font = section_font
ws_enum[f'A{start_fields_row}'].alignment = left_align
ws_enum[f'A{start_fields_row}'].fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
ws_enum[f'A{start_fields_row}'].border = thin_border

# 表头
header_row = start_fields_row + 1
for col_idx, h in enumerate(enum_headers, 1):
    cell = ws_enum.cell(row=header_row, column=col_idx)
    cell.value = h
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# 数据行
for row_idx, field in enumerate(enum_fields):
    r = header_row + 1 + row_idx
    for col_idx, val in enumerate(field, 1):
        cell = ws_enum.cell(row=r, column=col_idx)
        cell.value = val
        cell.font = normal_font
        cell.alignment = left_align if col_idx > 3 else center_align
        cell.border = thin_border
        if row_idx % 2 == 1:
            cell.fill = alt_fill

# 列宽
for i, w in enumerate(enum_col_widths, 1):
    ws_enum.column_dimensions[get_column_letter(i)].width = w

# 冻结窗格
ws_enum.freeze_panes = f'A{header_row + 1}'


# ========== 保存 ==========
output_path = '/Users/wangshuaijia/workspace/tyc/数据字典_powerlink.xlsx'
wb.save(output_path)
print(f"数据字典已生成: {output_path}")